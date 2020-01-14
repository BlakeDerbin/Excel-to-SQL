# -*- coding: utf-8 -*-
"""
Created on Wed Sep 18 07:48:22 2019

@author: Blake
"""

import openpyxl
import sqlite3
import os, sys

# Checks if the violation.db files exists and deletes it before starting the excel_data_import function
def check_db_file():
    db_file = "./violations.db"
    # If violations.db exists, delete it
    if os.path.isfile(db_file):        
        print("violations.db file already exists, deleting file.")
    
        # Removes database if it exits
        os.remove(db_file)
    else:
        excel_data_import()

# Checks whether the excel files are in the directory of this script, if not an error message is printed
def excel_data_import():
    try:
        print("Attempting to load excel workbooks, please wait...")

        # Loading the workbooks
        wb_inspections = openpyxl.load_workbook("inspections.xlsx")
        wb_violations = openpyxl.load_workbook("violations.xlsx")

        # loading the worksheet from excel
        ws_i = wb_inspections.active
        ws_v = wb_violations.active

        # Arrays for exporting excel data into
        inspections_data = []
        violations_data = []

        print("Starting import of excel workbooks, please wait...")

        # loops through worksheets and stores data in arrays
        for row in ws_i.iter_rows(min_row=2, max_col=20, values_only=True):
            inspections_data.append(row)

        for row in ws_v.iter_rows(min_row=2, max_col=5, values_only=True):
            violations_data.append(row)

        print("Excel import complete. Importing data into database, please wait...")

        create_sql_db(inspections_data, violations_data)

    except:
        print(
            "\nError: Please ensure the files inspections.xlsx and violations.xlsx exist in the same directory as the createdb_food.py script.\n")

        # Exits the script
        sys.exit

# Creates the sql database and imports the data from the excel arrays into their corresponding tables
def create_sql_db(array_1, array_2):
    try:
        # Creating/connecting to the database
        connection = sqlite3.connect("violations.db")
        cursor = connection.cursor()

        # SQL queries for creating inspections and violations tables
        create_inspections_table = """
            CREATE TABLE inspections (
            activity_date DATE,
            employee_id CHAR(9),
            facility_address VARCHAR(255),
            facility_city VARCHAR(255),
            facility_id VARCHAR(9),
            facility_name VARCHAR(500),
            facility_state CHAR(2),
            facility_zip VARCHAR(10),
            grade CHAR(1),
            owner_id CHAR(9),
            owner_name VARCHAR(50),
            pe_description VARCHAR(1000),
            program_element_pe INT(4),
            program_name VARCHAR(500),
            program_status VARCHAR(8),
            record_id CHAR(9),
            score INT(3),
            serial_number VARCHAR(9),
            service_code INT(3),
            service_description VARCHAR(50),
            PRIMARY KEY (serial_number)
            );"""

        create_violations_table = """
            CREATE TABLE violations (
            points INT(2),
            serial_number VARCHAR(9),
            violation_code CHAR(4),
            violation_description VARCHAR(1000),
            violation_status VARCHAR(50),
            FOREIGN KEY (serial_number) REFERENCES inspections(serial_number)
            );"""

        # Executing the SQL commands to create the tables
        cursor.execute(create_inspections_table)
        cursor.execute(create_violations_table)

        # Inserting the array data into the SQL database
        for d in array_1:
            cursor.execute('INSERT INTO inspections VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)', d)

        for d in array_2:
            cursor.execute('INSERT INTO violations VALUES (?,?,?,?,?)', d)

        # Committing the database changes and closing the connection
        connection.commit()
        connection.close()

        print("Data import complete.")
    except:
        print("Unknown error has occured")

check_db_file()